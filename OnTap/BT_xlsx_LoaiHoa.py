#Thư viện để ghi dữ liệu, format, tạo bảng biểu cho Excel 2010 (xlsx)
import xlsxwriter

'''List là dãy mà có khả năng lưu giữ các kiểu dữ liệu khác nhau và có thể thay đổi. Cho phép thành viên trùng lặp.
Tuple là một dãy giống List nhưng các đối tượng không thay đổi. Cho phép thành viên trùng lặp.
Set là một dãy không theo thứ tự và không được lập chỉ mục. Không chứa thành viên trùng lặp.
Dictionary là một dãy không theo thứ tự, có thể thay đổi và được lập chỉ mục.Không chứa thành viên trùng lặp.'''

'''
data=(['STT', 'Tên', 'Nơi trồng', 'Số lượng', 'Giá', 'Tổng'],
    [1,'Hoa Lài', 'Đà Lạt', 100, 10, ''],
    [2,'Hoa Cúc', 'Cần Thơ', 820, 15, ''],
    [3, 'Hoa Lan', 'Long An', 475, 30, ''],
    [4, 'Hoa Huệ', 'Tiền Giang', 624, 45, ''],
    [5, 'Hoa Giấy', 'Bến Tre', 752, 78, ''],
    [6, 'Hoa Mai', 'Vĩnh Long', 478, 92, ''],
    [7, 'Hoa Tulip', 'Đà Lạt', 354, 47, ''],)
#'''
#----------------------------------------------------------------------
#a. Trực quan hoá dữ liệu theo sơ đồ hình: cột, hình tròn (tính phần trăm) theo sl hoa

"""Ghi chú:
    **bar thì nhãn label
    **pie thì nhãn labels
=>> nếu k sẽ bị lỗi"""

    #Biểu đồ cột
"""
import matplotlib.pyplot as plt

data={'Hoa Lài':100,
    'Hoa Cúc':820,
    'Hoa Lan':475,
    'Hoa Huệ':624,
    'Hoa Giấy':752,
    'Hoa Mai':478,
    'Hoa Tulip':354}

#plt.bar(truc x, truc y)
plt.bar(range(len(data)), data.values(), label="Số lượng hoa")
plt.title("Biểu đồ tròn về số lượng hoa")
plt.xlabel("Tên hoa")
plt.ylabel("Số lượng")
plt.legend()
plt.xticks(range(len(data)),data.keys())
plt.show()
#"""
#----------------------------------------------------------------------

"""
    #Biểu đồ tròn
import matplotlib.pyplot as plt

data={'Hoa Lài':100,
    'Hoa Cúc':820,
    'Hoa Lan':475,
    'Hoa Huệ':624,
    'Hoa Giấy':752,
    'Hoa Mai':478,
    'Hoa Tulip':354}

plt.pie(data.values(),labels=data.keys(),autopct='%.2f%%')#auto percentage
plt.title("Biểu đồ tròn về số lượng hoa")
plt.show()
#"""
#----------------------------------------------------------------------

"""
#b.Tính tổng giá tiền các loại hoa
data=([1,'Hoa Lài', 'Đà Lạt', 100, 10, ''],
    [2,'Hoa Cúc', 'Cần Thơ', 820, 15, ''],
    [3, 'Hoa Lan', 'Long An', 475, 30, ''],
    [4, 'Hoa Huệ', 'Tiền Giang', 624, 45, ''],
    [5, 'Hoa Giấy', 'Bến Tre', 752, 78, ''],
    [6, 'Hoa Mai', 'Vĩnh Long', 478, 92, ''],
    [7, 'Hoa Tulip', 'Đà Lạt', 354, 47, ''],)

print("\nTổng giá tiền các loại hoa")
tong=[]
for i in range(len(data)):
    tong.append(data[i][3]*data[i][4])
    print(data[i][1],"\t",tong[i])
#"""

#----------------------------------------------------------------------

""""
#c. Lưu bảng dữ liệu xuống 1 file tên là xyz.xlsx
#row: hàng
#column: cột

import xlsxwriter

wB = xlsxwriter.Workbook("xyz.xlsx");
wS = wB.add_worksheet('Sheet 2')

data=(['STT', 'Tên', 'Nơi trồng', 'Số lượng', 'Giá', 'Tổng'],
    [1,'Hoa Lài', 'Đà Lạt', 100, 10, ''],
    [2,'Hoa Cúc', 'Cần Thơ', 820, 15, ''],
    [3, 'Hoa Lan', 'Long An', 475, 30, ''],
    [4, 'Hoa Huệ', 'Tiền Giang', 624, 45, ''],
    [5, 'Hoa Giấy', 'Bến Tre', 752, 78, ''],
    [6, 'Hoa Mai', 'Vĩnh Long', 478, 92, ''],
    [7, 'Hoa Tulip', 'Đà Lạt', 354, 47, ''],)

row, column = 0,0

#thêm dữ liệu vào trang tính
for sTT, ten, noiTrong, soLuong, donGia, tong in data:
    wS.write(row, column, sTT)
    wS.write(row, column + 1, ten)
    wS.write(row, column + 2, noiTrong)
    wS.write(row, column + 3, soLuong)
    wS.write(row, column + 4, donGia)
    wS.write(row, column + 5, tong)
    row += 1
wB.close()
#"""
#----------------------------------------------------------------------
"""
#d. Đọc dữ liệu file xyz.xlsx và in ra màn hình
import pandas as pd
df=pd.read_excel('xyz.xlsx')
print(df)
#"""

#----------------------------------------------------------------------
"""
#e. Cột tổng ở câu b lưu xuống file xyz
#thêm dữ liệu vào file xlsx
#Thư viện được đề xuất cho việc đọc ghi file Excel 2010 (xlsx)
import openpyxl

def getValueExcel(fileName, cellName):
    wB=openpyxl.load_workbook(fileName)
    wS=wB['Sheet 2']
    wB.close()
    return wS[cellName].value

def updateValueExcel(fileName, cellName, value):
    wB=openpyxl.load_workbook(fileName)
    wS=wB['Sheet 2']
    wS[cellName].value=value
    wB.close()
    wB.save(fileName)

xl = "xyz.xlsx"
tongList = []
#range(0,n-1)
for i in range(2,8+1): #range chạy từ 2-8
    tongList.append(getValueExcel(xl,'D'+str(i))*getValueExcel(xl,'E'+str(i)))
print(len(tongList))

#len(k đếm 0 )
for i in range(len(tongList)): #i: 0->7 =>> 2->8
    updateValueExcel(xl, 'F'+str(i + 2), tongList[i])

import pandas as pd
print(pd.read_excel(xl))

# """

#----------------------------------------------------------------------

#"""
#f. Tìm loại hoa có giá đắt nhất, số lượng ít nhất
import openpyxl

def getValueExcel(fileName, cellName):
    wB=openpyxl.load_workbook(fileName)
    wS=wB['Sheet 2']
    wB.close()
    return wS[cellName].value

xl = "xyz.xlsx"

    #Tìm max
max = getValueExcel(xl, 'E' + str(2))
viTri = 2

for i in range(2, 8 + 1):
    if (max < getValueExcel(xl, 'E' + str(i))):
        max = getValueExcel(xl, 'E' + str(i))
        viTri = i
print('Hoa đắt nhất là: ' + getValueExcel(xl, 'B' + str(viTri)), 'có đơn giá', max)

    #Tìm min
min = getValueExcel(xl, 'D' + str(2))
viTri2 = 2

for i in range(2, 8 + 1):
    if (min > getValueExcel(xl, 'D' + str(i))):
        min = getValueExcel(xl, 'D' + str(i))
        viTri2 = i
print('Hoa có số lượng ít nhất là: ' + getValueExcel(xl, 'B'+str(viTri2)), 'có số lượng', min)
#"""
