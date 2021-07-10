import pandas as pd
import matplotlib.pyplot as plt

data = ([1,"Hoa Lài","Đà Lạt",100,10],
        [2,"Hoa Cúc","Cần Thơ",820,15],
        [3,"Hoa Lan","Long An",475,30],
        [4,"Hoa Huệ","Tiền Giang",624,45],
        [5,"Hoa Giấy","Bến Tre",752,78],
        [6,"Hoa Mai","Vĩnh Long",478,92],
        [2,"Hoa Tulip","Đà Lạt",354,47],)
#a. Trực quan hoá dữ liệu theo sơ đồ hình: cột, hình tròn (tính phần trăm) theo sl hoa
name=[]
amount=[]
for i in range(len(data)):
    name.append(data[i][1])
    amount.append(data[i][3])


"""
Ghi chú:
    **bar thì nhãn label
    **pie thì nhãn labels
=>> nếu k sẽ bị lỗi
"""
    #Biểu đồ cột
"""
plt.bar(range(len(data)), amount, label="Số lượng hoa")
plt.title("Biểu đồ tròn về số lượng hoa")
plt.xlabel("Tên hoa")
plt.ylabel("Số lượng")
plt.xticks(range(len(data)),name)
plt.show()
#"""

"""
    #Biểu đồ tròn
plt.pie(amount,labels=name,autopct='%.2f%%')#auto percentage
plt.title("Biểu đồ tròn về số lượng hoa")
plt.show()
#"""

"""
#b.Tính tổng giá tiền các loại hoa
print("\nTổng giá tiền các loại hoa")
tong=[]
for i in range(len(data)):
    tong.append(data[i][3]*data[i][4])
    print(data[i][1],"\t",tong[i])
"""
#"""
#c. Lưu bảng dữ liệu xuống 1 file tên là xyz.xlsx
"""
#Tự làm
workbook.save("xyz.xlsx")
df=pd.DataFrame(data, columns=['STT','Tên','Nơi trồng','Số lượng','Đơn giá'])
df.to_excel(xl, index = False)
"""

"""
#Code thầy
import xlsxwriter

wB = xlsxwriter.Workbook("xyz.xlsx");
trang_tinh_1 = wB.add_worksheet('Sheet1')

duLieu=(['STT', 'Tên', 'Nơi trồng', 'Số lượng', 'Giá', 'Tổng'],
    [1,'Hoa Lài', 'Đà Lạt', 100, 10, ''],
    [2,'Hoa Cúc', 'Cần Thơ', 820, 15, ''],
    [3, 'Hoa Lan', 'Long An', 475, 30, ''],
    [4, 'Hoa Huệ', 'Tiền Giang', 624, 45, ''],
    [5, 'Hoa Giấy', 'Bến Tre', 752, 78, ''],
    [6, 'Hoa Mai', 'Vĩnh Long', 478, 92, ''],
    [7, 'Hoa Tulip', 'Đà Lạt', 354, 47, ''],)

row, column =0,0;
for sTT, ten, noiTrong, soLuong, donGia, tong in (duLieu):
    trang_tinh_1.write(row, column, sTT)
    trang_tinh_1.write(row, column + 1, ten)
    trang_tinh_1.write(row, column + 2, noiTrong)
    trang_tinh_1.write(row, column + 3, soLuong)
    trang_tinh_1.write(row, column + 4, donGia)
    trang_tinh_1.write(row, column + 5, tong)
    row += 1

wB.close()
#"""

"""
#d. Đọc dữ liệu file xyz.xlsx và in ra màn hình
df=pd.read_excel(xl)
print(df)
"""

#"""
#e. Cột tổng ở câu b lưu xuống file xyz
"""
#Câu gợi ý của thầy
import openpyxl

def getValueExcel(fileName, cellName):
    wB=openpyxl.load_workbook(fileName)
    trang_tinh_1=wB['Trang tính 1']
    wB.close()
    return trang_tinh_1[cellName].value

def updateValueExcel(fileName, cellName, value):
    wB=openpyxl.load_workbook(fileName)
    trang_tinh_1=wB['trang_tinh_1']
    trang_tinh_1[cellName].value=value
    wB.close()
    wB.save(fileName)

fileName = "xyz.xlsx"

tongList = []

for i in range(2, 8+1):
    tongList.append(getValueExcel(fileName,'D'+str(i))*getValueExcel(fileName,'E'+str(i)))
for i in range(len(tongList)):
    updateValueExcel(fileName, 'F'+str(i + 2), tongList[i])
"""

"""
df=pd.read_excel(xl)
df['Tổng']=df['Số lượng']*df['Đơn giá']
df.to_excel(xl, index= False)
print(df)
"""

#""
#f. Tìm loại hoa có giá đắt nhất, số lượng ít nhất
import openpyxl
def getValueExcel(fileName, cellName):
    wB = openpyxl.load_workbook(fileName)
    wS = wB['Trang tính 1']
    wB.close()
    return trang_tinh_1[cellName].valuefileName = "D:\\xyz.xlsx"
#Tìm max
max = getValueExcel(fileName, 'E' + str(2))
viTri = 0
for i in range(2, 8 + 1):
    if (max < getValueExcel(fileName, 'E' + str(i))):
        max = getValueExcel(fileName, 'E' + str(i))
        viTri = iprint('Hoa đắt nhất là: ' + getValueExcel(fileName, 'B' + str(viTri)), '-', max)
#Tìm min
min = getValueExcel(fileName, 'F' + str(2))
viTri2 = 0
for i in range(2, 8 + 1):
    if (min > getValueExcel(fileName, 'D' + str(i))):
        min = getValueExcel(fileName, 'D' + str(i))
        viTri2 = iprint('Hoa có số lượng ít nhất là: ' + getValueExcel(fileName, 'B' + str(viTri2)), '-', min)
#"""
